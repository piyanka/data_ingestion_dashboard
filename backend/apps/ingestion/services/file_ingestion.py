"""File parsing and persistence entrypoint.

This is intentionally thin. The assignment's real complexity is not in a clever
framework layer, but in keeping the raw payload, normalized record, validation
signals, and downstream emission calculation separate.
"""

from pathlib import Path
from typing import Iterable

from django.db import transaction

from apps.core.models import AuditLog

from ..models import RawRecord, SourceFile
from .normalization import normalize_raw_record
from .validation import validate_normalized_activity


def ingest_source_file(source_file: SourceFile) -> SourceFile:
    """Process the uploaded file synchronously for the prototype.

    A production version would hand this work to a background worker, but for
    a hiring exercise we want behavior that is easy to inspect end-to-end.
    """

    source_file.processing_status = SourceFile.ProcessingStatus.PROCESSING
    source_file.save(update_fields=["processing_status"])

    total_rows = successful_rows = failed_rows = 0

    try:
        for row_number, payload in enumerate(_iter_source_rows(source_file.uploaded_file), start=1):
            total_rows += 1
            raw_record = RawRecord.objects.create(
                source_file=source_file,
                row_number=row_number,
                raw_payload=payload,
                parse_status=RawRecord.ParseStatus.PARSED,
            )
            try:
                with transaction.atomic():
                    activity = normalize_raw_record(raw_record)
                    validate_normalized_activity(activity)
                    successful_rows += 1
            except Exception as exc:  # keep raw row even when normalization fails
                failed_rows += 1
                raw_record.parse_status = RawRecord.ParseStatus.FAILED
                raw_record.error_message = str(exc)
                raw_record.save(update_fields=["parse_status", "error_message"])
        source_file.total_rows = total_rows
        source_file.successful_rows = successful_rows
        source_file.failed_rows = failed_rows
        source_file.processing_status = (
            SourceFile.ProcessingStatus.PROCESSED if failed_rows == 0 else SourceFile.ProcessingStatus.PROCESSED_WITH_ERRORS
        )
        source_file.save(
            update_fields=["total_rows", "successful_rows", "failed_rows", "processing_status"]
        )
    except Exception as exc:
        source_file.processing_status = SourceFile.ProcessingStatus.FAILED
        source_file.save(update_fields=["processing_status"])
        raise exc

    AuditLog.objects.create(
        entity_type="SourceFile",
        entity_id=str(source_file.id),
        action_type="ingest",
        new_values={
            "total_rows": total_rows,
            "successful_rows": successful_rows,
            "failed_rows": failed_rows,
        },
    )
    return source_file


def _iter_source_rows(uploaded_file) -> Iterable[dict]:
    """Placeholder row reader.

    The concrete CSV/XLSX mapping will be source-specific because SAP, utility,
    and travel exports all disagree on column names and row structure.
    """

    path = Path(uploaded_file.name)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        import csv

        uploaded_file.open("r")
        try:
            yield from csv.DictReader(uploaded_file.file)
        finally:
            uploaded_file.close()
        return

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to read Excel uploads") from exc

        uploaded_file.open("rb")
        try:
            workbook = load_workbook(uploaded_file.file, data_only=True, read_only=True)
        finally:
            uploaded_file.close()
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return
        headers = [str(col).strip() if col is not None else "" for col in rows[0]]
        for row in rows[1:]:
            yield {headers[idx]: value for idx, value in enumerate(row)}
        return

    raise ValueError(f"Unsupported file type: {suffix}")
